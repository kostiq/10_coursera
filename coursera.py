import requests
import random
import re
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_list_of_random_courses(count=20):
    courses_xml = requests.get(
        'https://www.coursera.org/sitemap~www~courses.xml').content
    tree = etree.fromstring(courses_xml)
    courses_list = [link.find('{*}loc').text for link in tree]
    return random.sample(courses_list, count)


def get_title(html, soup):
    return (soup.find('h1').text)


def get_start_date(html, soup):
    if soup.find('div', {'class': re.compile('startdate')}):
        return soup.find('div', {'class': re.compile('startdate')}).text


def get_duration(html, soup):
    duration = ''
    class_list = soup.find_all('td', {'class': 'td-data'})
    if re.search('\d+-\d+', class_list[0].text):
        duration = class_list[0].text
    elif re.search('\d+-\d+', class_list[1].text):
        duration = class_list[1].text
    return duration


def get_avg_star(html, soup):
    if soup.find('div', {'class': re.compile('ratings-text')}):
        return soup.find('div', {'class': re.compile('ratings-text')}).text


def get_language(html, soup):
    return soup.find('div', {'class': 'language-info'}).text


def get_course_info(course_slug):
    course_info = []
    html = requests.get(course_slug).content
    soup = BeautifulSoup(html, 'html.parser')
    course_info = {'link': course_slug,
                   'title': get_title(html, soup),
                   'start_date': get_start_date(html, soup),
                   'duration': get_duration(html, soup),
                   'language': get_language(html, soup),
                   'avg_star': get_avg_star(html, soup)
                   }

    return course_info


def write_to_sheet(sheet, courses_info):
    ordered_course_fields = (
        'link', 'title', 'start_date', 'duration', 'language', 'avg_star')
    for x, course_info in enumerate(courses_info, start=1):
        for y, info in enumerate(course_info, start=1):
            sheet.cell(
                row=x,
                column=y,
                value=course_info[str(ordered_course_fields[y - 1])]
            )


def output_courses_info_to_xlsx(filepath, courses_info):
    wb = Workbook()
    sheet = wb.create_sheet(title='Coursera')
    write_to_sheet(sheet, courses_info)
    wb.save(filename=filepath)


if __name__ == '__main__':
    output_courses_info_to_xlsx(
        'info.xlsx',
        [get_course_info(link) for link in get_list_of_random_courses()]
    )
